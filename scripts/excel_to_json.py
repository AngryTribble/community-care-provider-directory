"""
Community Care Compass - Excel to JSON Converter

Purpose:
Reads the master provider workbook and generates data/providers.json for the website.

Expected repo structure:
  source/provider_master.xlsx
  scripts/excel_to_json.py
  data/providers.json

Run from the repository root:
  python scripts/excel_to_json.py

Install dependency if needed:
  pip install openpyxl
"""

from __future__ import annotations

import json
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


# ---------- SETTINGS ----------

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "source" / "provider_master.xlsx"
OUTPUT_PATH = ROOT / "providers.json"  # FIX: was Path("proivders.json") — typo + wrong location

# Tabs that should not be treated as provider specialty tabs.
IGNORE_SHEETS = {
    "README",
    "Instructions",
    "Dropdown_Lists",
    "Change_Log",
    "Change_Requests",
}

# PPMS column names, exactly as you described them.
PPMS_COLUMNS = {
    "network_name": "Network Name",
    "status": "Status",
    "vss": "VSS",
    "availability": "Availability",
    "provider": "Provider",
    "provider_notes": "Notes (Provider) (Provider)",
    "provider_identifier": "Provider Identifier",
    "hpp": "HPP",
    "specialty": "Specialty",
    "specialty_group": "Specialty Group",
    "subservice_rollup": "Subservice Rollup",
    "telehealth_available": "Telehealth Available",
    "phone": "Care Site Phone Number",
    "fax": "Care Site Fax (Caresite) (Care Site)",
    "caresite": "Caresite",
    "street": "Care Site Address",
    "city": "Care Site City",
    "state": "Care Site State",
    "zip": "Care Site Zip Code",
    "organization_group": "Organization / Group",
    "created_on": "Created On",
    "deactivation_date": "Deactivation Date",
    "coverage_county": "Coverage County",
}

# Compass-added columns.
COMPASS_COLUMNS = {
    "display": "Compass_Display",
    "include": "Compass_Include",
    "office_nickname": "Compass_Office_Nickname",
    "referral_fax": "Compass_Referral_Fax",
    "medical_records_fax": "Compass_Medical_Records_Fax",
    "email": "Compass_Email",
    "hsrm_status": "Compass_HSRM_Status",
    "eps_status": "Compass_EPS_Status",
    "accepting_status": "Compass_Accepting_Status",
    "preferred_method": "Compass_Preferred_Method",
    "internal_notes": "Compass_Internal_Notes",
    "referral_comment_notes": "Compass_Referral_Comment_Notes",
    "tags": "Compass_Tags",
    "last_verified": "Compass_Last_Verified",
    "verified_by": "Compass_Verified_By",
    "update_status": "Compass_Update_Status",
    "update_notes": "Compass_Update_Notes",
    "latitude": "Latitude",
    "longitude": "Longitude",
    "geocode_status": "Geocode_Status",
}


# ---------- HELPER FUNCTIONS ----------

def clean(value: Any) -> str:
    """Convert Excel cell value to a clean string."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def split_list(value: Any) -> List[str]:
    """Split tags/notes separated by semicolon, pipe, comma, or new lines."""
    text = clean(value)
    if not text:
        return []
    parts = re.split(r"[;|\n]+", text)
    return [p.strip() for p in parts if p.strip()]


def slugify(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"(^-|-$)", "", text)
    return text or "provider"


def yes_no_unknown(value: Any, default: str = "Unknown") -> str:
    text = clean(value)
    if not text:
        return default
    low = text.lower()
    if low in {"yes", "y", "true", "active", "available"}:
        return "Yes"
    if low in {"no", "n", "false", "inactive", "not available"}:
        return "No"
    return text


def header_map(sheet) -> Dict[str, int]:
    """Return mapping of header name to 1-based column index."""
    headers = {}
    for cell in sheet[1]:
        name = clean(cell.value)
        if name:
            headers[name] = cell.column
    return headers


def get(row, headers: Dict[str, int], column_name: str) -> str:
    col = headers.get(column_name)
    if not col:
        return ""
    return clean(row[col - 1].value)


def should_include(row, headers: Dict[str, int]) -> bool:
    include_value = get(row, headers, COMPASS_COLUMNS["include"]).lower()
    deactivation = get(row, headers, PPMS_COLUMNS["deactivation_date"])
    status = get(row, headers, PPMS_COLUMNS["status"]).lower()

    if include_value in {"no", "n", "false", "exclude", "0"}:
        return False
    if deactivation and include_value not in {"yes", "y", "true", "include", "1"}:
        return False
    if status in {"inactive", "terminated", "deactivated"} and include_value not in {"yes", "y", "true", "include", "1"}:
        return False
    return True


def build_provider(row, headers: Dict[str, int], sheet_name: str, used_ids: set[str]) -> Dict[str, Any]:
    ppms = {key: get(row, headers, col) for key, col in PPMS_COLUMNS.items()}
    compass = {key: get(row, headers, col) for key, col in COMPASS_COLUMNS.items()}

    office_name = compass["office_nickname"] or ppms["caresite"] or ppms["organization_group"] or ppms["provider"] or "Unknown Provider"
    provider_name = ppms["provider"]
    city = ppms["city"]
    state = ppms["state"]
    zip_code = ppms["zip"]

    base_id = slugify(f"{office_name}-{provider_name}-{city}-{state}-{zip_code}")
    provider_id = base_id
    counter = 2
    while provider_id in used_ids:
        provider_id = f"{base_id}-{counter}"
        counter += 1
    used_ids.add(provider_id)

    referral_fax = compass["referral_fax"] or ppms["fax"]

    notes = []
    notes.extend(split_list(ppms["provider_notes"]))
    notes.extend(split_list(compass["internal_notes"]))
    notes.extend(split_list(compass["referral_comment_notes"]))

    specialties = [ppms["specialty"] or sheet_name]
    if ppms["specialty_group"] and ppms["specialty_group"] not in specialties:
        specialties.append(ppms["specialty_group"])

    tags = split_list(compass["tags"])
    if ppms["subservice_rollup"]:
        tags.append(ppms["subservice_rollup"])
    if ppms["telehealth_available"]:
        tags.append(f"Telehealth: {ppms['telehealth_available']}")

    return {
        "id": provider_id,
        "sourceSheet": sheet_name,
        "officeName": office_name,
        "providerName": provider_name,
        "specialties": specialties,
        "tags": tags,
        "address": {
            "street": ppms["street"],
            "city": city,
            "state": state,
            "zip": zip_code,
        },
        "phone": ppms["phone"],
        "fax": referral_fax,
        "medicalRecordsFax": compass["medical_records_fax"],
        "email": compass["email"],
        "providerNpi": ppms["provider_identifier"],
        "npi": ppms["provider_identifier"],
        "careSiteNpi": "",
        "optumStatus": ppms["status"] or "Unknown",
        "optumStatus": ppms["status"] or "Unknown",
        "availability": ppms["availability"],
        "telehealthAvailable": ppms["telehealth_available"],
        "epsStatus": yes_no_unknown(compass["eps_status"]),
        "hsrmStatus": compass["hsrm_status"] or "Unknown",
        "acceptingNewReferrals": compass["accepting_status"] or ppms["availability"] or "Unknown",
        "preferredReferralMethod": compass["preferred_method"] or "Fax",
        "coverageCounty": ppms["coverage_county"],
        "organizationGroup": ppms["organization_group"],
        "networkName": ppms["network_name"],
        "vss": ppms["vss"],
        "hpp": ppms["hpp"],
        "createdOn": ppms["created_on"],
        "deactivationDate": ppms["deactivation_date"],
        "notes": notes,
        "restrictions": [],
        "lastVerified": compass["last_verified"],
        "verifiedBy": compass["verified_by"],
        "updateStatus": compass["update_status"],
        "updateNotes": compass["update_notes"],
        "latitude": compass["latitude"],
        "longitude": compass["longitude"],
        "geocodeStatus": compass["geocode_status"],
        "display": compass["display"],
    }


# ---------- MAIN ----------

def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    providers: List[Dict[str, Any]] = []
    used_ids: set[str] = set()

    for sheet_name in wb.sheetnames:
        if sheet_name in IGNORE_SHEETS:
            continue

        sheet = wb[sheet_name]
        headers = header_map(sheet)

        # Skip sheets that do not look like PPMS provider tabs.
        required = [
            PPMS_COLUMNS["provider"],
            PPMS_COLUMNS["specialty"],
            PPMS_COLUMNS["caresite"],
            PPMS_COLUMNS["street"],
        ]
        if not all(col in headers for col in required):
            print(f"Skipping sheet '{sheet_name}' because required PPMS columns were not found.")
            continue

        sheet_count = 0
        for row in sheet.iter_rows(min_row=2):
            provider = get(row, headers, PPMS_COLUMNS["provider"])
            caresite = get(row, headers, PPMS_COLUMNS["caresite"])
            street = get(row, headers, PPMS_COLUMNS["street"])

            if not any([provider, caresite, street]):
                continue
            if not should_include(row, headers):
                continue

            providers.append(build_provider(row, headers, sheet_name, used_ids))
            sheet_count += 1

        print(f"Processed {sheet_count} providers from '{sheet_name}'.")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_PATH.open("w", encoding="utf-8") as f:
        json.dump(providers, f, indent=2, ensure_ascii=False)

    print(f"\nDone. Wrote {len(providers)} providers to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
